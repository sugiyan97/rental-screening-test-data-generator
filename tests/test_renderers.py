import csv
import json
import zipfile

import pytest

from rental_pdf_generator.generator import CasePdfGenerator
from rental_pdf_generator.models import Case, DocumentSpec
from rental_pdf_generator.renderers import (
    OUTPUT_FORMATS,
    UnsupportedOutputFormatError,
    render_document,
)

# 出力形式ごとの先頭バイト（形式が実際にその種類になっているかの検証用）
MAGIC_BYTES = {
    "pdf": b"%PDF",
    "png": b"\x89PNG",
    "jpg": b"\xff\xd8\xff",
    "xlsx": b"PK",
    "docx": b"PK",
    "pptx": b"PK",
}


def _single_document_case(base_case: Case, output_format: str) -> Case:
    case = base_case.model_copy(deep=True)
    case.case_id = f"CASE-FMT-{output_format.upper()}"
    case.documents = [
        DocumentSpec(
            document_type="rental_application_individual",
            variant="standard",
            output_format=output_format,
        )
    ]
    return case


def test_default_output_format_is_pdf():
    spec = DocumentSpec(document_type="income_certificate", variant="salary_certificate")
    assert spec.output_format == "pdf"


def test_unknown_output_format_is_rejected():
    with pytest.raises(ValueError):
        DocumentSpec(
            document_type="income_certificate",
            variant="salary_certificate",
            output_format="tiff",
        )


def test_render_document_unsupported_format_raises(tmp_path):
    with pytest.raises(UnsupportedOutputFormatError):
        render_document(page=None, output_format="tiff", path=tmp_path / "x.tiff")


@pytest.mark.parametrize("output_format", OUTPUT_FORMATS)
def test_generate_each_output_format(individual_case, tmp_path, output_format):
    case = _single_document_case(individual_case, output_format)
    generator = CasePdfGenerator(output_dir=tmp_path)
    meta = generator.generate(case)

    expected = f"{output_format}/rental_application_individual_standard.{output_format}"
    doc_path = tmp_path / case.case_id / expected
    assert doc_path.exists(), f"{output_format} が生成されていない"
    assert doc_path.stat().st_size > 0

    if output_format in MAGIC_BYTES:
        with doc_path.open("rb") as f:
            assert f.read(len(MAGIC_BYTES[output_format])) == MAGIC_BYTES[output_format]

    entry = meta["generated_documents"][0]
    assert entry["output_format"] == output_format
    assert entry["file"] == expected
    # 正解 JSON は形式に関わらず同じパスに出力される
    answers_dir = tmp_path / case.case_id / "answers"
    assert (answers_dir / "rental_application_individual_standard.json").exists()


def test_pdf_entry_keeps_pdf_key_for_compatibility(individual_case, tmp_path):
    case = _single_document_case(individual_case, "pdf")
    meta = CasePdfGenerator(output_dir=tmp_path).generate(case)
    entry = meta["generated_documents"][0]
    assert entry["pdf"] == entry["file"] == "pdf/rental_application_individual_standard.pdf"


def test_non_pdf_entry_has_no_pdf_key(individual_case, tmp_path):
    case = _single_document_case(individual_case, "docx")
    meta = CasePdfGenerator(output_dir=tmp_path).generate(case)
    assert "pdf" not in meta["generated_documents"][0]


def test_csv_contains_rendered_values(individual_case, tmp_path):
    case = _single_document_case(individual_case, "csv")
    CasePdfGenerator(output_dir=tmp_path).generate(case)

    csv_path = tmp_path / case.case_id / "csv" / "rental_application_individual_standard.csv"
    with csv_path.open(encoding="utf-8", newline="") as f:
        cells = [cell for row in csv.reader(f) for cell in row]
    joined = " ".join(cells)
    assert "テスト 花子" in joined
    assert "テストアパート" in joined


def test_office_formats_are_valid_zip_packages(individual_case, tmp_path):
    generator = CasePdfGenerator(output_dir=tmp_path)
    for output_format, member in (
        ("xlsx", "xl/workbook.xml"),
        ("docx", "word/document.xml"),
        ("pptx", "ppt/presentation.xml"),
    ):
        case = _single_document_case(individual_case, output_format)
        generator.generate(case)
        path = (
            tmp_path
            / case.case_id
            / output_format
            / f"rental_application_individual_standard.{output_format}"
        )
        with zipfile.ZipFile(path) as zf:
            assert member in zf.namelist()


def test_docx_contains_rendered_values(individual_case, tmp_path):
    case = _single_document_case(individual_case, "docx")
    CasePdfGenerator(output_dir=tmp_path).generate(case)

    path = tmp_path / case.case_id / "docx" / "rental_application_individual_standard.docx"
    with zipfile.ZipFile(path) as zf:
        xml = zf.read("word/document.xml").decode("utf-8")
    assert "テスト 花子" in xml


def test_xlsx_contains_rendered_values(individual_case, tmp_path):
    from openpyxl import load_workbook

    case = _single_document_case(individual_case, "xlsx")
    CasePdfGenerator(output_dir=tmp_path).generate(case)

    path = tmp_path / case.case_id / "xlsx" / "rental_application_individual_standard.xlsx"
    sheet = load_workbook(path).active
    values = [str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value]
    assert any("テスト 花子" in v for v in values)


def test_output_format_override_applies_to_all_documents(individual_case, tmp_path):
    generator = CasePdfGenerator(output_dir=tmp_path, output_format="png")
    meta = generator.generate(individual_case)

    assert all(d["output_format"] == "png" for d in meta["generated_documents"])
    assert not (tmp_path / individual_case.case_id / "pdf").exists()
    for doc in meta["generated_documents"]:
        assert (tmp_path / individual_case.case_id / doc["file"]).exists()


def test_payment_track_record_pledge_generated_as_pdf_and_docx(
    corporate_extended_case, tmp_path
):
    """同一の支払実績確約書を PDF と Word(.docx) の2形式で出力できる（TC-1-102）。"""
    case = corporate_extended_case.model_copy(deep=True)
    case.case_id = "CASE-PLEDGE-MULTI"
    case.documents = [
        DocumentSpec(
            document_type="payment_track_record_pledge",
            variant="standard",
            output_format="pdf",
        ),
        DocumentSpec(
            document_type="payment_track_record_pledge",
            variant="standard",
            output_format="docx",
        ),
    ]
    meta = CasePdfGenerator(output_dir=tmp_path).generate(case)

    case_dir = tmp_path / case.case_id
    pdf_path = case_dir / "pdf" / "payment_track_record_pledge_standard.pdf"
    docx_path = case_dir / "docx" / "payment_track_record_pledge_standard.docx"
    assert pdf_path.exists()
    assert docx_path.exists()
    with pdf_path.open("rb") as f:
        assert f.read(4) == b"%PDF"

    with zipfile.ZipFile(docx_path) as zf:
        xml = zf.read("word/document.xml").decode("utf-8")
    # 確約者・延滞履歴・完済状況が Word 本文（表）に入っている
    assert "テスト商事株式会社" in xml
    assert "延滞なし" in xml
    assert "完済（未払残高 0円）" in xml

    assert [d["output_format"] for d in meta["generated_documents"]] == ["pdf", "docx"]


def test_case_meta_json_records_output_format(individual_case, tmp_path):
    case = _single_document_case(individual_case, "jpg")
    CasePdfGenerator(output_dir=tmp_path).generate(case)

    meta = json.loads(
        (tmp_path / case.case_id / "case_meta.json").read_text(encoding="utf-8")
    )
    assert meta["generated_documents"][0]["output_format"] == "jpg"
